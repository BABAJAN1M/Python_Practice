import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    

    cell_a1 = sheet['A1']
    cell_1_1 = sheet.cell(1, 1)

    print(cell_a1.value)


    max_rows = sheet.max_row
    print("Max Rows:", max_rows)
    print('---')

    for row in range(2, max_rows + 1):
     
        cell = sheet.cell(row, 3)
        print("Original Price (Row {}): {}".format(row, cell.value))

        corrected_price = cell.value * 0.9

        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

   
    values = Reference(sheet, min_row=2, max_row=max_rows, min_col=4, max_col=4)

  
    chart = BarChart()
    chart.add_data(values)


    sheet.add_chart(chart, 'E2')


    wb.save('transact2.xlsx')

process_workbook('transact.xlsx')
